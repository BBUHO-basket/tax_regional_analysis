import openpyxl
import pandas as pd
from pathlib import Path

BASE_DIR    = Path(__file__).parent.parent
FILE_PATH   = BASE_DIR / 'data' / 'raw' / '2025년_2기_부가가치세_신고현황.xlsx'
OUTPUT_PATH = BASE_DIR / 'data' / 'processed' / 'vat_processed.csv'

INDUSTRIES = [
    '농업·임업·어업',
    '광업',
    '제조업',
    '전기·가스·수도업',
    '도매업',
    '소매업',
    '부동산임대업',
    '건설업',
    '음식업',
    '숙박업',
    '운수창고통신업 등',
    '부동산임대업 외',
    '대리중개도급 등',
    '기타서비스업',
]


def parse_num(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s in ('-', ''):
        return 0.0
    if s.startswith('- '):
        s = '-' + s[2:]
    s = s.replace(',', '')
    try:
        return float(s)
    except ValueError:
        return None


def parse_multiline(val):
    if val is None:
        return [None] * 14
    items = str(val).split('\n')
    result = [parse_num(item) for item in items]
    while len(result) < 14:
        result.append(None)
    return result[:14]


def extract_records(ws):
    records = []

    # Format A: 서울(row 5), 인천(row 20) — 개별 행 1업종씩
    for total_row in [5, 20]:
        region = ws.cell(total_row, 1).value.strip()
        for i, industry in enumerate(INDUSTRIES):
            r = total_row + 1 + i
            records.append({
                'region':           region,
                'industry':         industry,
                'tax_people_count': parse_num(ws.cell(r, 4).value),
                'tax_base':         parse_num(ws.cell(r, 8).value),
                'zero_tax_base':    parse_num(ws.cell(r, 17).value),
                'pay_tax':          parse_num(ws.cell(r, 19).value),
                'refund_tax':       parse_num(ws.cell(r, 21).value),
            })

    # Format B: 경기 등 13개 지역 — 줄바꿈으로 묶인 멀티라인 셀
    for total_row in [38, 58, 114, 134, 157, 177, 200, 220, 243, 263, 286, 306, 329]:
        region = ws.cell(total_row, 1).value.strip()
        dr = total_row + 1
        e = parse_multiline(ws.cell(dr, 5).value)
        j = parse_multiline(ws.cell(dr, 10).value)
        q = parse_multiline(ws.cell(dr, 17).value)
        s = parse_multiline(ws.cell(dr, 19).value)
        v = parse_multiline(ws.cell(dr, 22).value)
        for i, industry in enumerate(INDUSTRIES):
            records.append({
                'region':           region,
                'industry':         industry,
                'tax_people_count': e[i],
                'tax_base':         j[i],
                'zero_tax_base':    q[i],
                'pay_tax':          s[i],
                'refund_tax':       v[i],
            })

    # Format C: 대전(81), 충북(96) — 개별 행, 열 위치 다름
    for total_row in [81, 96]:
        region = ws.cell(total_row, 1).value.strip()
        for i, industry in enumerate(INDUSTRIES):
            r = total_row + 1 + i
            records.append({
                'region':           region,
                'industry':         industry,
                'tax_people_count': parse_num(ws.cell(r, 4).value),
                'tax_base':         parse_num(ws.cell(r, 9).value),
                'zero_tax_base':    parse_num(ws.cell(r, 16).value),
                'pay_tax':          parse_num(ws.cell(r, 18).value),
                'refund_tax':       parse_num(ws.cell(r, 20).value),
            })

    return records


def main():
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active

    records = extract_records(ws)

    df = pd.DataFrame(records)
    df['final_tax'] = df['pay_tax'] - df['refund_tax']

    for col in ['tax_people_count', 'tax_base', 'zero_tax_base', 'pay_tax', 'refund_tax', 'final_tax']:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    df.to_csv(OUTPUT_PATH, index=False, encoding='utf-8-sig')

    print(f'저장 완료: {OUTPUT_PATH}')
    print(f'총 레코드: {len(df)}행  ({df["region"].nunique()}개 지역 × {len(INDUSTRIES)}개 업종)')
    print(f'결측값:\n{df.isnull().sum()}')


if __name__ == '__main__':
    main()
